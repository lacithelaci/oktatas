from openpyxl.workbook import Workbook
from src.data.basic.model_dataclasses import User, Playlist, Song


def write_users_xlsx(users: list[User], wb: Workbook,
                 sheet_name: str = "users",
                 heading: bool = True) -> None:
    sheet = wb.create_sheet(sheet_name)

    if heading:
        column_names = ["id", "name", "age", "country"]
        for col in range(len(column_names)):
            sheet.cell(row=1, column=col + 1, value=column_names[col])

    offset = 2 if heading else 1
    for row in range(len(users)):
        sheet.cell(row=row + offset, column=1, value=users[row].id)
        sheet.cell(row=row + offset, column=2, value=users[row].name)
        sheet.cell(row=row + offset, column=3, value=users[row].age)
        sheet.cell(row=row + offset, column=4, value=users[row].country)


def read_users_xlsx(wb: Workbook,
                sheet_name: str = "users",
                heading: bool = True) -> list[User]:
    sheet = wb[sheet_name]

    users = []
    row = 2 if heading else 1
    while True:
        cell = sheet.cell(row=row, column=1)
        if cell.value is None:
            break

        users.append(
            User(
                sheet.cell(row=row, column=1).value,
                sheet.cell(row=row, column=2).value,
                sheet.cell(row=row, column=3).value,
                sheet.cell(row=row, column=4).value
            )
        )
        row += 1
    return users


def write_playlists_xlsx(playlists: list[Playlist], wb: Workbook,
                 sheet_name: str = "playlists",
                 heading: bool = True) -> None:
    sheet = wb.create_sheet(sheet_name)

    if heading:
        column_names = ["id", "owner_id", "name", "songs", "private"]
        for col in range(len(column_names)):
            sheet.cell(row=1, column=col + 1, value=column_names[col])

    offset = 2 if heading else 1
    for row in range(len(playlists)):
        sheet.cell(row=row + offset, column=1, value=playlists[row].id)
        sheet.cell(row=row + offset, column=2, value=playlists[row].owner_id)
        sheet.cell(row=row + offset, column=3, value=playlists[row].name)
        sheet.cell(row=row + offset, column=4, value=playlists[row].songs)
        sheet.cell(row=row + offset, column=5, value=playlists[row].private)


def read_playlists_xlsx(wb: Workbook,
                sheet_name: str = "playlists",
                heading: bool = True) -> list[Playlist]:
    sheet = wb[sheet_name]

    playlists = []
    row = 2 if heading else 1
    while True:
        cell = sheet.cell(row=row, column=1)
        if cell.value is None:
            break

        playlists.append(
            Playlist(
                sheet.cell(row=row, column=1).value,
                sheet.cell(row=row, column=2).value,
                sheet.cell(row=row, column=3).value,
                sheet.cell(row=row, column=4).value,
                sheet.cell(row=row, column=5).value
            )
        )
        row += 1
    return playlists


def write_songs_xlsx(songs: list[Song], wb: Workbook,
                 sheet_name: str = "songs",
                 heading: bool = True) -> None:
    sheet = wb.create_sheet(sheet_name)

    if heading:
        column_names = ["id", "playlist_id", "title", "artist", "genre"]
        for col in range(len(column_names)):
            sheet.cell(row=1, column=col + 1, value=column_names[col])

    offset = 2 if heading else 1
    for row in range(len(songs)):
        sheet.cell(row=row + offset, column=1, value=songs[row].id)
        sheet.cell(row=row + offset, column=2, value=songs[row].playlist_id)
        sheet.cell(row=row + offset, column=3, value=songs[row].title)
        sheet.cell(row=row + offset, column=4, value=songs[row].artist)
        sheet.cell(row=row + offset, column=5, value=songs[row].genre)


def read_songs_xlsx(wb: Workbook,
                sheet_name: str = "songs",
                heading: bool = True) -> list[Song]:
    sheet = wb[sheet_name]

    songs = []
    row = 2 if heading else 1
    while True:
        cell = sheet.cell(row=row, column=1)
        if cell.value is None:
            break

        songs.append(
            Song(
                sheet.cell(row=row, column=1).value,
                sheet.cell(row=row, column=2).value,
                sheet.cell(row=row, column=3).value,
                sheet.cell(row=row, column=4).value,
                sheet.cell(row=row, column=5).value
            )
        )
        row += 1
    return songs
